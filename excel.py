from openpyxl import load_workbook

def excel():
    wb = load_workbook('D:/Users/Jrodriguezc/OneDrive - Linea Directa S.A.S/Escritorio/Scripting/PlantillaCargaHilos.xlsx')
    sheet = wb ["Hoja1"]
    

    for i in range (2,sheet.max_row):
        if sheet.cell(row=i,column=1).value==None:
            return i 
        print(i)




#----------------------------------------------------------------------------------------------------------------------------------------

# filesheet = "D:/Users/Jrodriguezc/OneDrive - Linea Directa S.A.S/Escritorio/Scripting/PlantillaCargaHilos.xlsx"

# wb = load_workbook(filesheet)

# sheet = wb.active

# sheet['A7'] = 32

# sheet['B7'] = 1845

# wb.save(filesheet)




#---------------------------------------------------------------------------------------------------------------------------------------

# filesheet = "D:/Users/Jrodriguezc/OneDrive - Linea Directa S.A.S/Escritorio/Scripting/PlantillaCargaHilos.xlsx"

# wb = load_workbook(filesheet)

# sheet = wb.active

# C1 = sheet['C1'].value

# C2 = sheet['C2'].value

# C3 = sheet['C3'].value

# celdas = [C1, C2, C3]
# for valor in celdas:
#  print(valor)




#-----------------------------------------------------------------------------------------------------------------------------------------------------------

# wb = openpyxl.load_workbook("D:/Users/Jrodriguezc/OneDrive - Linea Directa S.A.S/Escritorio/Scripting/PlantillaCargaHilos.xlsx", read_only=True)
# ws = wb.active
 
# for row in ws.rows:
#     for cell in row:
#         print(cell.value)


# wb = openpyxl.Workbook(write_only=True)
# ws = wb.create_sheet()
 
# for row in range(10):
#     ws.append([f"{i}" for i in range(20)])
 
# wb.save("D:/Users/Jrodriguezc/OneDrive - Linea Directa S.A.S/Escritorio/Scripting/PlantillaCargaHilos.xlsx")




#------------------------------------------------------------------------------------------------------------------------------

# from openpyxl import load_workbook


# filesheet = "D:/Users/Jrodriguezc/OneDrive - Linea Directa S.A.S/Escritorio/Scripting/PlantillaCargaHilos.xlsx"

# # creamos ell obejeto load_workbook
# wb = load_workbook(filesheet)

# # seleccionamos el archivo
# sheet = wb.active

# # Obtenemos el valor de la celda A1
# A1 = sheet['A1'].value

# B1 = sheet['B1'].value

# C1 = sheet['C1'].value

# D1 = sheet['D1'].value

# E1 = sheet['E1'].value

# G1 = sheet['G1'].value


# # Mostramos los valores 
# celdas = [A1, B1, C1, D1, E1, G1]
# for valor in celdas:
#     print(valor)



#--------------------------------------------------------------------------------------------------------------------------------------------------

# import pandas as pd

# def excel():
#     archivo_excel = pd.read_excel('D:/Users/Jrodriguezc/OneDrive - Linea Directa S.A.S/Escritorio/Scripting/PlantillaCargaHilos.xlsx')
#     print(archivo_excel.columns)
#     values = archivo_excel['Material'].values

#     print(values)
    
#     columnas = ['Material', 'METROS HILO PT', 'METROS NYLON PT','COSTO HILO PT','COSTO NYLON PT','METROS TOTALES PT','COSTO TOTAL PT']
#     df_seleccionados = archivo_excel[columnas]


#     print(df_seleccionados)


#     return df_seleccionados 


# fila = session.findById("wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MNAME[0,4]").text 

#     print(fila)
#     for fila in range (9):
#         valor = session.findById(f"wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MNAME[0,{fila}]").text
#         print(valor)