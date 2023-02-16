import win32com.client as win32
from openpyxl import load_workbook
from log_clase import log
from datetime import datetime
import sys

# conxion con sap

def conexion():

    SapGuiAuto = win32.GetObject("SAPGUI")
    if not type(SapGuiAuto) == win32.CDispatch:
        return
    application = SapGuiAuto.GetScriptingEngine
    if not type(application) == win32.CDispatch:
        SapGuiAuto = None
        return
    connection = application.Children(0)
    if not type(connection) == win32.CDispatch:
        application = None
        SapGuiAuto = None
        return
    session = connection.Children(0)
    if not type(session) == win32.CDispatch:
        connection = None
        application = None
        SapGuiAuto = None
        return
    return session

# lectura del excel y la funcion principal de ejecucion


def ingresar_material(ruta_archivo):
    try:

        # conxion con excel y libreria del archivo
        wb = load_workbook(ruta_archivo)

        sheet = wb["Hoja1"]  # buscando dentro del archivo la hoja

        for i in range(2, sheet.max_row + 1):    # recorre la hoja
            # valida que el valor de la fila sea un valor valido
            if sheet.cell(row=i, column=1).value == None:
                break

            rn = datetime.now()
            logs.idtransaccion = datetime.now().strftime("%H:%M:%S-")+str(rn.microsecond)
            logs.fechainicio = datetime.now()
            logs.level = 'INFO'
            logs.procesointerno = 'Inicio Transaccion'
            logs.mensaje = 'Inicio Transaccion'
            logs.post()

            session.findById("wnd[0]").maximize()
            session.findById("wnd[0]/tbar[0]/okcd").text = "/nmm02"

            logs.level = 'INFO'
            logs.procesointerno = 'mm02'
            logs.mensaje = 'Ingresando Valores'
            logs.post()

            session.findById("wnd[0]").sendVKey(0)
            # ingresand el valor del archivo excel dentro de las vistas en sap
            session.findById(
                "wnd[0]/usr/ctxtRMMG1-MATNR").text = sheet.cell(row=i, column=1).value
            session.findById("wnd[0]/usr/ctxtRMMG1-MATNR").caretPosition = 6
            session.findById("wnd[0]").sendVKey(0)
            session.findById(
                "wnd[1]/usr/tblSAPLMGMMTC_VIEW").getAbsoluteRow(0).selected = True
            session.findById("wnd[1]/tbar[0]/btn[0]").press()

            session.findById(
                "wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S").verticalScrollbar.position = 6
            session.findById(
                "wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,2]").text = sheet.cell(row=i, column=2).value
            session.findById(
                "wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,3]").text = sheet.cell(row=i, column=3).value
            session.findById(
                "wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,4]").text = sheet.cell(row=i, column=4).value
            session.findById(
                "wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,5]").text = sheet.cell(row=i, column=5).value
            session.findById(
                "wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,6]").text = sheet.cell(row=i, column=6).value
            session.findById(
                "wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,7]").text = sheet.cell(row=i, column=7).value
            session.findById("wnd[0]/tbar[0]/btn[11]").press()

            # da respuesta a materiales modificados en el archivo excel
            sheet.cell(row=i, column=8).value = session.findById(
                "wnd[0]/sbar").text
            wb.save(ruta_archivo)  # guardar archivo excel

            logs.level = 'INFO'
            logs.procesointerno = 'Fin Transaccion'
            logs.mensaje = 'Fin Transaccion'
            logs.fintransaccion = datetime.now()
            logs.postFin()

            print(session.findById("wnd[0]/sbar").text)

    except Exception as e:

        print(e)

# adminstra y llama a las funciones


def main(clasificacion_material):
    global logs
    logs = log()
    global session
    session = conexion()
    ingresar_material(clasificacion_material)
