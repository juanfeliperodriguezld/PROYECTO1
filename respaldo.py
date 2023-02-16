import win32com.client as win32
from excel import excel 
from openpyxl import load_workbook

# conxion con sap

def conexion (): 
   
    SapGuiAuto = win32.GetObject("SAPGUI")
    if not type(SapGuiAuto) == win32.CDispatch:
        return
    application = SapGuiAuto.GetScriptingEngine
    if not type(application) == win32.CDispatch:
        SapGuiAuto = None
        return
    connection = application.Children(0)
    if not type(connection) == win32.CDispatch:
        application = None
        SapGuiAuto = None
        return
    session = connection.Children(0)
    if not type(session) == win32.CDispatch:
        connection = None
        application = None
        SapGuiAuto = None
        return
    return session

# lectura del excel y la funcion principal de ejecucion 

def ingresar_material(ruta_archivo):
   
   wb = load_workbook(ruta_archivo)  # conxion con excel y libreria del archivo

   sheet = wb ["Hoja1"] # buscando dentro del archivo la hoja 

   for i in range (2,sheet.max_row +1):    # recorre la hoja 
    if sheet.cell(row=i,column=1).value==None:  # valida que el valor de la fila sea un valor valido 
        break

    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nmm02"

    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/usr/ctxtRMMG1-MATNR").text = sheet.cell(row=i,column=1).value  # ingresand el valor del archivo excel dentro de las vistas en sap 
    session.findById("wnd[0]/usr/ctxtRMMG1-MATNR").caretPosition = 6
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[1]/usr/tblSAPLMGMMTC_VIEW").getAbsoluteRow(0).selected = True
    session.findById("wnd[1]/tbar[0]/btn[0]").press()

    session.findById("wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S").verticalScrollbar.position = 6   
    session.findById("wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,2]").text = sheet.cell(row=i,column=2).value
    session.findById("wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,3]").text = sheet.cell(row=i,column=3).value
    session.findById("wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,4]").text = sheet.cell(row=i,column=4).value
    session.findById("wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,5]").text = sheet.cell(row=i,column=5).value
    session.findById("wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,6]").text = sheet.cell(row=i,column=6).value
    session.findById("wnd[0]/usr/subSUBSCR_BEWERT:SAPLCTMS:5000/tabsTABSTRIP_CHAR/tabpTAB1/ssubTABSTRIP_CHAR_GR:SAPLCTMS:5100/tblSAPLCTMSCHARS_S/ctxtRCTMS-MWERT[1,7]").text = sheet.cell(row=i,column=7).value
    session.findById("wnd[0]/tbar[0]/btn[11]").press()
    
    sheet.cell(row=i,column=8).value = session.findById("wnd[0]/sbar").text  #  da respuesta a materiales modificados en el archivo excel
    wb.save(ruta_archivo)  # guardar archivo excel 

    print (session.findById("wnd[0]/sbar").text)


# adminstra y llama a las funciones 

if __name__ == "__main__":
    global session
    session = conexion()
    ingresar_material('D:/Users/Jrodriguezc/OneDrive - Linea Directa S.A.S/Escritorio/Scripting/PlantillaCargaHilos.xlsx')
